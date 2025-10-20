import math
import os
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse, HttpResponseBadRequest
from rest_framework import generics, views, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.authtoken.models import Token
from django.contrib import messages
from .forms import HotelForm
from .forms import MenuItemForm
from .models import Category, Hotel, ManualCategory, MenuItem
from .serializers import HotelSerializer, HotelDetailSerializer, MenuItemSerializer
from .utils import download_image
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from collections import defaultdict
from slugify import slugify
from django.contrib.auth import authenticate, login, logout
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from rest_framework.pagination import LimitOffsetPagination
from django.db.models import Count
from random import sample
import random
from rest_framework.pagination import PageNumberPagination
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from rest_framework.generics import ListAPIView

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, "login.html")

@login_required
def logout_view(request):
    logout(request)
    return redirect("login")

@login_required
def dashboard_view(request):
    hotels = Hotel.objects.all().order_by('-uploaded_at')
    return render(request, "dashboard.html", {"hotels": hotels})

def hotel_list_view(request):
    """List all hotels uploaded."""
    hotels = Hotel.objects.all().order_by('-uploaded_at')
    return render(request, "hotel_list.html", {"hotels": hotels})

def hotel_menu_view(request, pk):
    hotel = get_object_or_404(Hotel, pk=pk)
    menu_items = MenuItem.objects.filter(hotel=hotel).prefetch_related('categories', 'manual_categories')
    grouped = defaultdict(list)
    for item in menu_items:
        cats = list(item.categories.all())
        cat = cats[0].name if cats else "Uncategorized"
        grouped[cat].append(item)

    context = {
        "hotel": hotel,
        "grouped_menu": dict(sorted(grouped.items())),
        "all_categories": Category.objects.all(),
        "all_manual_categories": ManualCategory.objects.all(),
    }
    return render(request, "hotel_menu.html", context)

@login_required
def upload_xlsx_view(request):
    """Upload .xlsx file and import data into DB (supports _shortDescEn and _hotelLogo)."""
    if request.method == "POST":
        xlsx_file = request.FILES.get("file")
        if not xlsx_file:
            messages.error(request, "Please upload a file.")
            return redirect("upload")

        if not xlsx_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are allowed.")
            return redirect("upload")

        # Extract hotel name from file name
        filename = xlsx_file.name
        if "menu_data_" in filename:
            hotel_name = filename.split("menu_data_")[-1].replace(".xlsx", "")
        else:
            hotel_name = filename.replace(".xlsx", "")
        hotel_name = hotel_name.strip()

        hotel, _ = Hotel.objects.get_or_create(name=hotel_name)

        # Save file temporarily
        tmp_dir = os.path.join(settings.MEDIA_ROOT, "temp")
        os.makedirs(tmp_dir, exist_ok=True)
        tmp_path = os.path.join(tmp_dir, xlsx_file.name)

        with open(tmp_path, "wb+") as dest:
            for chunk in xlsx_file.chunks():
                dest.write(chunk)

        # Load workbook
        try:
            wb = load_workbook(tmp_path, data_only=True)
        except Exception:
            try:
                df = pd.read_excel(tmp_path, engine="openpyxl")
                df.to_excel(tmp_path, index=False, engine="openpyxl")
                wb = load_workbook(tmp_path, data_only=True)
            except Exception:
                messages.error(request, "Invalid Excel file. Please re-save properly.")
                return redirect("upload")

        sheet = wb.active
        imported = 0

        # Read headers (lowercased)
        headers = [str(cell.value).strip().lower() for cell in sheet[1] if cell.value]

        def find_col(possible_names):
            for name in possible_names:
                if name in headers:
                    return headers.index(name)
            return None

        # Match your Excel headers
        idx_name = find_col(["_nameen", "name", "item", "itemname"])
        idx_price = find_col(["_finalprice", "_regularprice", "price", "regularprice"])
        idx_category = find_col(["_categories", "category", "type"])
        idx_description = find_col(["_shortdescen", "description", "desc", "details"])
        idx_image = find_col(["_imageurl1", "image", "image_url", "imageurl"])
        idx_logo = find_col(["_hotellogo", "logo", "hotel_logo"])

        # ✅ Handle hotel logo (first row only)
        if idx_logo is not None:
            try:
                logo_val = sheet.cell(row=2, column=idx_logo + 1).value
                if logo_val:
                    logo_url = str(logo_val).strip()
                    logo_local = download_image(logo_url, slugify(hotel_name) + "_logo")
                    if logo_local:
                        hotel.logo = logo_local
                        hotel.save()
            except Exception as e:
                print(f"[Hotel Logo Error] {e}")

        # Clear existing menu items for hotel
        MenuItem.objects.filter(hotel=hotel).delete()

        # ✅ Iterate rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                name = row[idx_name] if idx_name is not None else None
                if not name:
                    continue
                name = str(name).strip()

                # Price
                price = None
                if idx_price is not None and row[idx_price] is not None:
                    try:
                        price = float(row[idx_price])
                    except:
                        price = None

                # Category
                categories = []
                if idx_category is not None and row[idx_category] is not None:
                    cat_cell = str(row[idx_category]).strip()
                    if cat_cell:
                        cat_names = [c.strip() for c in cat_cell.split(",") if c.strip()]
                        for cat_name in cat_names:
                            cat_obj, _ = Category.objects.get_or_create(name__iexact=cat_name, defaults={'name': cat_name})
                            categories.append(cat_obj)

                # Description
                description = None
                if idx_description is not None and row[idx_description] is not None:
                    description = str(row[idx_description]).strip()

                # Image
                image_url = None
                if idx_image is not None and row[idx_image] is not None:
                    image_url = str(row[idx_image]).strip()

                image_local = None
                if image_url:
                    image_local = download_image(image_url, slugify(hotel_name))

                item = MenuItem.objects.create(
                    hotel=hotel,
                    item_name=name,
                    price=price,
                    description=description,
                    image_url=image_url,
                    image_local=image_local,
                )
                if categories:
                    item.categories.set(categories)  # ✅ link many-to-many

            except Exception as e:
                print(f"[Row Error] {e}")

        wb.close()
        os.remove(tmp_path)

        messages.success(request, f"{imported} menu items imported for {hotel_name}.")
        return redirect("dashboard")

    return render(request, "upload.html")

@login_required
def add_hotel_view(request):
    """Add a new hotel manually"""
    if request.method == 'POST':
        form = HotelForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Hotel added successfully!")
            return redirect('dashboard')
    else:
        form = HotelForm()
    return render(request, 'hotel_form.html', {'form': form, 'title': 'Add Hotel'})


@login_required
def edit_hotel_view(request, pk):
    """Edit an existing hotel"""
    hotel = get_object_or_404(Hotel, pk=pk)
    if request.method == 'POST':
        form = HotelForm(request.POST, request.FILES, instance=hotel)
        if form.is_valid():
            form.save()
            messages.success(request, "Hotel updated successfully!")
            return redirect('dashboard')
    else:
        form = HotelForm(instance=hotel)
    return render(request, 'hotel_form.html', {'form': form, 'title': f'Edit {hotel.name}'})

@login_required
def add_menu_item_view(request, hotel_id):
    hotel = get_object_or_404(Hotel, pk=hotel_id)

    if request.method == 'POST':
        form = MenuItemForm(request.POST, request.FILES)
        if form.is_valid():
            menu_item = form.save(commit=False)
            menu_item.hotel = hotel

            upload_file = form.cleaned_data.get('image_upload')
            image_url = form.cleaned_data.get('image_url')

            # ✅ If a file was uploaded, use it
            if upload_file:
                menu_item.image_local = upload_file
                menu_item.image_url = ""
            # ✅ If URL given, download it locally
            elif image_url:
                local_path = download_image(image_url, hotel.name)
                if local_path:
                    menu_item.image_local = local_path
                    menu_item.image_url = image_url

            menu_item.save()
            form.save_m2m()

            messages.success(request, f"Menu item '{menu_item.item_name}' added successfully.")
            return redirect('hotel_menu', pk=hotel_id)
    else:
        form = MenuItemForm()

    return render(request, 'menu_form.html', {'form': form, 'hotel': hotel, 'title': f"Add Menu Item - {hotel.name}"})


@login_required
def edit_menu_item_view(request, hotel_id, item_id):
    hotel = get_object_or_404(Hotel, pk=hotel_id)
    menu_item = get_object_or_404(MenuItem, pk=item_id, hotel=hotel)

    old_image_path = None
    if menu_item.image_local:
        old_image_path = os.path.join(settings.MEDIA_ROOT, str(menu_item.image_local))

    if request.method == 'POST':
        form = MenuItemForm(request.POST, request.FILES, instance=menu_item)
        if form.is_valid():
            menu_item = form.save(commit=False)

            upload_file = form.cleaned_data.get('image_upload')
            image_url = form.cleaned_data.get('image_url')

            # ✅ CASE 1: Uploaded file — delete old image and save new one
            if upload_file:
                if old_image_path and os.path.exists(old_image_path):
                    os.remove(old_image_path)
                menu_item.image_local = upload_file
                menu_item.image_url = ""

            # ✅ CASE 2: URL provided — download & replace old local image
            elif image_url:
                local_path = download_image(image_url, hotel.name)
                if local_path:
                    if old_image_path and os.path.exists(old_image_path):
                        os.remove(old_image_path)
                    menu_item.image_local = local_path
                    menu_item.image_url = image_url

            menu_item.save()
            form.save_m2m()

            messages.success(request, f"Menu item '{menu_item.item_name}' updated successfully.")
            return redirect('hotel_menu', pk=hotel_id)
    else:
        form = MenuItemForm(instance=menu_item)

    return render(request, 'menu_form.html', {
        'form': form,
        'hotel': hotel,
        'title': f"Edit Menu Item - {hotel.name}"
    })

# -----------------------------
# Web Upload View (login required)
# -----------------------------
# @login_required
# def upload_page(request):
#     if request.method == 'POST':
#         file = request.FILES.get('file')
#         if not file:
#             return render(request, 'upload.html', {'error': 'No file uploaded'})
#         if not file.name.lower().endswith('.xlsx'):
#             return render(request, 'upload.html', {'error': 'Only .xlsx files allowed'})


#         # Extract hotel name from filename
#         fname = os.path.splitext(file.name)[0]
#         # expected pattern: menu_data_<Hotel Name>
#         hotel_name = fname.replace('menu_data_', '').replace('menu-data-', '').strip()
#         if not hotel_name:
#             hotel_name = fname


#         # Save temporary file
#         tmp_path = os.path.join(settings.MEDIA_ROOT, 'uploads')
#         os.makedirs(tmp_path, exist_ok=True)
#         tmp_file_path = os.path.join(tmp_path, slugify(file.name))
#         with open(tmp_file_path, 'wb') as dest:
#             for chunk in file.chunks():
#                 dest.write(chunk)


#         try:
#             imported = import_xlsx_to_db(tmp_file_path, hotel_name)
#         except Exception as e:
#             return render(request, 'upload.html', {'error': f'Import failed: {e}'})


#         return render(request, 'upload.html', {'success': True, 'imported': imported, 'hotel_name': hotel_name})


#     return render(request, 'upload.html')


# -----------------------------
# XLSX parsing and import logic
# -----------------------------
@transaction.atomic
def import_xlsx_to_db(filepath: str, hotel_name: str) -> dict:
    """Parse an xlsx file and save Hotel and MenuItems. Returns dict with counts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    hotel, created = Hotel.objects.get_or_create(name=hotel_name)
    hotel.menu_items.all().delete()

    rows_processed = 0
    items_created = 0

    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]

    def col_index(key_names):
        for k in key_names:
            if k in headers:
                return headers.index(k)
        return None

    idx_item = col_index(['item', 'item_name', 'name'])
    idx_price = col_index(['price', 'cost'])
    idx_category = col_index(['category', 'type'])
    idx_description = col_index(['description', 'desc', 'details', 'item_description'])  # ✅
    idx_image = col_index(['image', 'image_url', 'image link', 'imageurl', 'image_link'])

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows_processed += 1
        try:
            item_name = row[idx_item] if idx_item is not None else None
            if not item_name:
                continue
            item_name = str(item_name).strip()

            price = None
            if idx_price is not None and row[idx_price] is not None:
                try:
                    price = float(row[idx_price])
                except Exception:
                    price = None

            category = str(row[idx_category]).strip() if idx_category is not None and row[idx_category] else None
            description = str(row[idx_description]).strip() if idx_description is not None and row[idx_description] else None
            image_url = str(row[idx_image]).strip() if idx_image is not None and row[idx_image] else None

            image_local_rel = ''
            if image_url:
                image_local_rel = download_image(image_url, hotel_name)

            obj, created_item = MenuItem.objects.get_or_create(
                hotel=hotel,
                item_name=item_name,
                defaults={
                    'price': price,
                    'category': category,
                    'description': description,
                    'image_url': image_url,
                    'image_local': image_local_rel,
                }
            )
            if not created_item:
                changed = False
                for field, new_val in {
                    'price': price,
                    'category': category,
                    'description': description,
                    'image_url': image_url,
                    'image_local': image_local_rel
                }.items():
                    if new_val and getattr(obj, field) != new_val:
                        setattr(obj, field, new_val)
                        changed = True
                if changed:
                    obj.save()
            else:
                items_created += 1
        except Exception as e:
            print(f"[Row Error] {e}")
            continue

    wb.close()
    return {'rows_processed': rows_processed, 'items_created': items_created}


# -----------------------------
# API: Upload (token or api key protected)
# -----------------------------





class UploadMenuAPI(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)


    def post(self, request, format=None):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.lower().endswith('.xlsx'):
            return Response({'detail': 'Only .xlsx allowed'}, status=status.HTTP_400_BAD_REQUEST)


        # Get hotel name from filename
        fname = os.path.splitext(file.name)[0]
        hotel_name = fname.replace('menu_data_', '').replace('menu-data-', '').strip()
        if not hotel_name:
            hotel_name = fname


        # Save temporary
        tmp_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        os.makedirs(tmp_dir, exist_ok=True)
        tmp_file = os.path.join(tmp_dir, slugify(file.name))
        with open(tmp_file, 'wb') as dest:
            for chunk in file.chunks():
                dest.write(chunk)


        result = import_xlsx_to_db(tmp_file, hotel_name)
        return Response({'hotel': hotel_name, 'items_imported': result.get('items_created', 0), 'status': 'success'})

# -----------------------------
# API: Mobile endpoints
# -----------------------------

class FifteenItemPagination(PageNumberPagination):
    page_size = 15
    page_size_query_param = 'page_size'
    max_page_size = 50

    def get_paginated_response(self, data):
        total_pages = math.ceil(self.page.paginator.count / self.page_size)
        next_link = self.get_next_link()
        prev_link = self.get_previous_link()

        return Response({
            'count': self.page.paginator.count,
            'total_pages': total_pages,
            'has_next': bool(next_link),       # ✅ boolean (True/False)
            'has_previous': bool(prev_link),   # ✅ boolean (True/False)
            'next': next_link,
            'previous': prev_link,
            'results': data
        })

class UnifiedMenuAPI(ListAPIView):
    permission_classes = [AllowAny]
    pagination_class = FifteenItemPagination

    def get_queryset(self):
        items = (
            MenuItem.objects.filter(is_visible=True)
            .select_related("hotel")
            .prefetch_related("manual_categories")
        )
        return items

    def list(self, request, *args, **kwargs):
        items = self.get_queryset()

        # ✅ Group items by item_name
        grouped = defaultdict(list)
        for item in items:
            grouped[item.item_name.strip()].append(item)

        # ✅ Build unique item entries with hotel list
        unique_items = []
        for item_name, menu_list in grouped.items():
            # pick first item to represent general menu details
            first_item = menu_list[0]

            hotel_data = []
            for m in menu_list:
                hotel_data.append({
                    "hotel_name": m.hotel.name,
                    "hotel_logo": request.build_absolute_uri(m.hotel.logo.url) if m.hotel.logo else None
                })

            unique_items.append({
                "item_name": first_item.item_name,
                "price": f"{float(first_item.price):.3f} KD" if first_item.price is not None else None,
                "description": first_item.description,
                "image": request.build_absolute_uri(first_item.image_url) if first_item.image_url else (
                    request.build_absolute_uri(f"/media/{first_item.image_local}") if first_item.image_local else None
                ),
                "manual_categories": [cat.name for cat in first_item.manual_categories.all()],
                "hotels": hotel_data
            })

        # ✅ Shuffle the unique items
        random.shuffle(unique_items)

        # ✅ Paginate
        page = self.paginate_queryset(unique_items)
        if page is not None:
            return self.get_paginated_response(page)

        return Response(unique_items)


class HotelListAPI(generics.ListAPIView):
    queryset = Hotel.objects.all().order_by('-uploaded_at')
    serializer_class = HotelSerializer
    permission_classes = (AllowAny,)




class HotelMenuAPI(generics.RetrieveAPIView):
    queryset = Hotel.objects.all()
    serializer_class = HotelDetailSerializer
    permission_classes = (AllowAny,)


class MenuAllAPI(generics.ListAPIView):
    serializer_class = MenuItemSerializer
    permission_classes = (AllowAny,)
    pagination_class = LimitOffsetPagination

    def get_queryset(self):
        return MenuItem.objects.select_related('hotel').filter(is_visible=True).order_by('item_name')


class MenuRandomAPI(views.APIView):
    permission_classes = (AllowAny,)

    def get(self, request, format=None):
        count = int(request.GET.get('count', 10))
        qs = list(MenuItem.objects.select_related('hotel').all())
        total = len(qs)
        if total == 0:
            return Response({'items': []})
        if count >= total:
            chosen = qs
        else:
            # random sample
            chosen = sample(qs, min(count, total))
        serializer = MenuItemSerializer(chosen, many=True, context={'request': request})
        return Response({'count': len(chosen), 'items': serializer.data})
    
    


class ToggleVisibilityAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            item = MenuItem.objects.get(pk=pk)
            visible = request.data.get("visible", True)
            item.is_visible = bool(visible)
            item.save()
            return Response({"status": "success", "visible": item.is_visible})
        except MenuItem.DoesNotExist:
            return Response({"error": "Menu item not found"}, status=status.HTTP_404_NOT_FOUND)
        
@require_POST
@csrf_exempt   # You can remove this if you prefer to use CSRF token in fetch
def create_manual_category(request):
    """Create a new manual category via AJAX."""
    name = request.POST.get("name", "").strip()
    if not name:
        return JsonResponse({"error": "Category name cannot be empty."}, status=400)

    category, created = ManualCategory.objects.get_or_create(name__iexact=name, defaults={"name": name})
    if created:
        return JsonResponse({"status": "success", "id": category.id, "name": category.name})
    else:
        return JsonResponse({"status": "exists", "id": category.id, "name": category.name})

class UpdateMenuCategoryAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        type_ = request.data.get("type")
        category_id = request.data.get("category_id")

        try:
            item = MenuItem.objects.get(pk=pk)
            if type_ == "auto":
                if category_id:
                    category = Category.objects.get(pk=category_id)
                    item.categories.set([category])
                else:
                    item.categories.clear()
            elif type_ == "manual":
                if category_id:
                    category = ManualCategory.objects.get(pk=category_id)
                    item.manual_categories.set([category])
                else:
                    item.manual_categories.clear()
            item.save()
            return Response({"status": "success"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)
        
class DeleteMenuItemAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            item = MenuItem.objects.get(pk=pk)
            # delete local image if exists
            if item.image_local:
                file_path = os.path.join(settings.MEDIA_ROOT, str(item.image_local))
                if os.path.exists(file_path):
                    os.remove(file_path)
            item.delete()
            return Response({"success": True})
        except MenuItem.DoesNotExist:
            return Response({"error": "Menu item not found"}, status=404)


class BulkDeleteMenuItemsAPI(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = request.data.get("ids", [])
        deleted = 0

        for pk in ids:
            try:
                item = MenuItem.objects.get(pk=pk)
                if item.image_local:
                    file_path = os.path.join(settings.MEDIA_ROOT, str(item.image_local))
                    if os.path.exists(file_path):
                        os.remove(file_path)
                item.delete()
                deleted += 1
            except MenuItem.DoesNotExist:
                continue

        return Response({"deleted": deleted})